"""
Čita popis polaznika iz Excel tablica tečajeva (generiranih zoho_to_excel.py)
radi povezivanja bankovnih uplata s pravom osobom, i radi upisivanja
uplaćenog iznosa natrag u tablicu.
"""

import unicodedata
from pathlib import Path

from openpyxl import load_workbook

FIRST_PARTICIPANT_ROW = 10


def strip_diacritics(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def normalize_name(text: str) -> str:
    return strip_diacritics(text).upper().strip()


def load_registrants(excel_dir: Path) -> list:
    """Vrati listu dictova za svakog polaznika u svim .xlsx datotekama u
    excel_dir (osim predloška): {first_name, last_name, name_variants,
    email, course_code, location, file_path, row}."""
    registrants = []

    for path in sorted(Path(excel_dir).glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue  # privremena Excel zaključana datoteka
        try:
            wb = load_workbook(path)
        except Exception:
            continue
        if "podaci" not in wb.sheetnames:
            continue
        ws = wb["podaci"]

        course_code = ws["C4"].value or ""
        location = ws["C5"].value or ""

        row = FIRST_PARTICIPANT_ROW
        while True:
            first_name = ws.cell(row=row, column=2).value
            last_name = ws.cell(row=row, column=3).value
            row_number = ws.cell(row=row, column=1).value
            if row_number is None:
                break  # izvan tablice polaznika (stigli do reda ukupnog iznosa)

            if first_name:
                full = f"{first_name} {last_name or ''}".strip()
                reversed_name = f"{last_name or ''} {first_name}".strip()
                registrants.append({
                    "first_name": str(first_name).strip(),
                    "last_name": str(last_name or "").strip(),
                    "name_variants": [normalize_name(full), normalize_name(reversed_name)],
                    "email": ws.cell(row=row, column=8).value or "",
                    "course_code": course_code,
                    "location": location,
                    "file_path": path,
                    "row": row,
                })
            row += 1

    return registrants


def find_matching_registrant(raw_line: str, registrants: list):
    """Traži bilo koju poznatu varijantu imena polaznika kao podniz u
    (normaliziranom) tekstu retka transakcije. Vraća prvi pronađeni
    registrant dict, ili None."""
    normalized_line = normalize_name(raw_line)
    for registrant in registrants:
        for variant in registrant["name_variants"]:
            if variant and variant in normalized_line:
                return registrant
    return None


def add_payment(file_path: Path, row: int, amount: float) -> float:
    """Doda amount na postojeću vrijednost u koloni 'Payment Received' (L)
    za taj red (akumulira ako već postoji npr. akontacija). Vrati novi
    ukupni iznos."""
    wb = load_workbook(file_path)
    ws = wb["podaci"]
    cell = ws.cell(row=row, column=12)
    existing = cell.value or 0
    try:
        existing = float(existing)
    except (TypeError, ValueError):
        existing = 0
    new_total = existing + amount
    cell.value = new_total
    wb.save(file_path)
    return new_total
