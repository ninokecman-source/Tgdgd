"""
Čita prijave (mailove) sa Zoho Mail računa i upisuje podatke polaznika u
Excel tablicu po uzoru na Emmett Technique Instructor Administration Sheet.

Za svaku kombinaciju (kod tečaja + grad održavanja) postoji zasebna .xlsx
datoteka, npr. "Modul 1&2 Split.xlsx". Nova prijava se dodaje kao novi red
u tablicu polaznika te datoteke.

Pokretanje:
    python zoho_to_excel.py

Postavke se čitaju iz config.json (napravi ga iz config.example.json).
"""

import imaplib
import smtplib
import email
from email import policy
from email.message import EmailMessage
from datetime import datetime
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

CONFIG_PATH = Path(__file__).with_name("config.json")

# --- Predložak: struktura Emmett administrativne tablice -------------------

FIRST_PARTICIPANT_ROW = 10
TOTALS_ROW = 29  # red s '=SUM(...)' formulom, pomiče se ako se doda red

PHONE_FORMAT = "@"

FIELD_LABELS = {
    "ime_prezime": "Ime i prezime",
    "email": "Email",
    "mjesto_postanski": "Mjesto stanovanja i poštanski broj",
    "ulica": "Adresa",
    "telefon": "Br. Tel",
}

POSTAL_CODE_RE = re.compile(r"\b\d{5}\b")
DATE_RE = re.compile(r"\d{1,2}\.\s*-?\s*\d{0,2}\.?\s*\d{1,2}\.\s*\d{4}\.?")
TAG_RE = re.compile(r"<[^<]+?>")


def load_config() -> dict:
    if not CONFIG_PATH.exists():
        sys.exit(
            f"Nedostaje {CONFIG_PATH}. Kopiraj config.example.json u config.json "
            "i popuni svoje podatke."
        )
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


def load_state(state_path: Path) -> set:
    if state_path.exists():
        with open(state_path, encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_state(state_path: Path, processed: set) -> None:
    with open(state_path, "w", encoding="utf-8") as f:
        json.dump(sorted(processed), f, ensure_ascii=False, indent=2)


def get_plain_text(msg) -> str:
    body = msg.get_body(preferencelist=("plain", "html"))
    if body is None:
        return ""
    content = body.get_content()
    if body.get_content_type() == "text/html":
        content = TAG_RE.sub("\n", content)
        content = (
            content.replace("&nbsp;", " ")
            .replace("&amp;", "&")
            .replace("&lt;", "<")
            .replace("&gt;", ">")
        )
    return content


def find_identifier_line(lines: list) -> str:
    for line in lines:
        if line.strip().lower().startswith("prijava"):
            continue
        if " - " in line or re.search(r"-\s*\d{2}[.\-]", line):
            return line
    return lines[0] if lines else ""


def extract_value_after_label(lines: list, label: str) -> str:
    for i, line in enumerate(lines):
        if line.strip().startswith(label):
            for next_line in lines[i + 1:]:
                if next_line.strip():
                    return next_line.strip()
    return ""


def split_mjesto_postanski(value: str) -> tuple:
    match = POSTAL_CODE_RE.search(value)
    if not match:
        return value.strip(), ""
    postal_code = match.group(0)
    grad = (value[: match.start()] + value[match.end():]).strip()
    return grad, postal_code


def split_first_last_name(full_name: str) -> tuple:
    parts = full_name.strip().split(maxsplit=1)
    if len(parts) == 2:
        return parts[0], parts[1]
    if len(parts) == 1:
        return parts[0], ""
    return "", ""


def parse_application(text: str) -> dict:
    lines = [ln.strip() for ln in text.splitlines()]
    non_empty = [ln for ln in lines if ln]

    identifier_line = find_identifier_line(non_empty)

    mjesto_raw = extract_value_after_label(lines, FIELD_LABELS["mjesto_postanski"])
    grad, postal_code = split_mjesto_postanski(mjesto_raw)

    full_name = extract_value_after_label(lines, FIELD_LABELS["ime_prezime"])
    first_name, last_name = split_first_last_name(full_name)

    return {
        "identifier_line": identifier_line,
        "First name": first_name,
        "Last Name": last_name,
        "Street ": extract_value_after_label(lines, FIELD_LABELS["ulica"]),
        "City/ Town ": grad,
        "Post/Zip Code ": postal_code,
        "Email Address": extract_value_after_label(lines, FIELD_LABELS["email"]),
        "Mobile Ph": extract_value_after_label(lines, FIELD_LABELS["telefon"]),
    }


def split_after_course_code(identifier_line: str, course_code: str):
    """Vrati (grad, datum_i_ostatak) - sve između koda tečaja i prve
    znamenke je grad (radi za bilo koji grad), a od prve znamenke nadalje
    je datum (+ eventualno ime instruktora na kraju). Ovo je otporno na
    razne zapise datuma koje EMMETT koristi u praksi (18/19.01.2025.,
    22-23.03.2025., 30.11.-01.12.2024. itd.) jer se ne oslanja na strogi
    regex za datum, nego samo traži gdje datum počinje."""
    code_match = re.search(re.escape(course_code), identifier_line, re.IGNORECASE)
    if not code_match:
        return "", ""
    rest = identifier_line[code_match.end():]
    digit_match = re.search(r"\d", rest)
    if not digit_match:
        return rest.strip(" -\t/"), ""
    location = rest[:digit_match.start()].strip(" -\t/")
    remainder = rest[digit_match.start():]
    return location, remainder


def extract_location(identifier_line: str, course_code: str) -> str:
    location, _ = split_after_course_code(identifier_line, course_code)
    return location


def extract_dates(identifier_line: str, course_code: str = None) -> str:
    if course_code is None:
        match = DATE_RE.search(identifier_line)
        return match.group(0) if match else ""
    _, remainder = split_after_course_code(identifier_line, course_code)
    return remainder.split(" - ", 1)[0].strip()


def extract_trailing_name(identifier_line: str, course_code: str) -> str:
    """Ime instruktora na kraju retka, ako ga mail sadrži (noviji format).
    Vraća prazan string ako nije prisutno (stariji mailovi)."""
    _, remainder = split_after_course_code(identifier_line, course_code)
    parts = remainder.split(" - ", 1)
    return parts[1].strip() if len(parts) == 2 else ""


def match_course_and_location(identifier_line: str, course_codes: list):
    """Kod tečaja mora biti jedan od poznatih (Modul 1&2, Modul 3, ...) -
    to je dovoljno da isključi mailove koji nisu prijave na tečaj (npr.
    generičke obavijesti). Ime instruktora se više ne provjerava jer ga
    stariji mailovi ne sadrže, a u praksi svi mailovi s poznatim kodom u
    ovom sandučiću i tako pripadaju tebi."""
    course_code = None
    for code in sorted(course_codes, key=len, reverse=True):
        if code.lower() in identifier_line.lower():
            course_code = code
            break

    if not course_code:
        return None, None

    location = extract_location(identifier_line, course_code)
    if not location:
        return None, None

    return course_code, location


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "", name).strip()


FOLDER_LIST_RE = re.compile(r'^\((?P<flags>[^)]*)\)\s+"(?P<delim>[^"]*)"\s+"(?P<name>.*)"$')


def imap_utf7_decode(s: str) -> str:
    """Dekodira IMAP naziv foldera iz modificiranog UTF-7 (RFC 3501) u
    normalan tekst, npr. 'Modul 1&-2' -> 'Modul 1&2',
    'Upla&AQc-eno' -> 'Uplaćeno'."""
    import base64

    res = []
    b64_chars = ""
    in_b64 = False
    for ch in s:
        if not in_b64:
            if ch == "&":
                in_b64 = True
                b64_chars = ""
            else:
                res.append(ch)
        else:
            if ch == "-":
                if b64_chars == "":
                    res.append("&")
                else:
                    padded = b64_chars.replace(",", "/")
                    padded += "=" * (-len(padded) % 4)
                    res.append(base64.b64decode(padded).decode("utf-16-be"))
                in_b64 = False
            else:
                b64_chars += ch
    if in_b64 and b64_chars:
        padded = b64_chars.replace(",", "/")
        padded += "=" * (-len(padded) % 4)
        res.append(base64.b64decode(padded).decode("utf-16-be"))
    return "".join(res)


def discover_target_folders(imap: imaplib.IMAP4_SSL, roots: list) -> list:
    """Vrati listu (sirovi_naziv_foldera, lokacija) za sve foldere koji
    pripadaju jednom od zadanih 'root' foldera (npr. 'Split', 'Zagreb') ili
    su njihovi podfolderi (npr. 'Zagreb/Modul 3'). 'Lokacija' je prvi dio
    puta (npr. 'Zagreb' za 'Zagreb/Modul 3'). Sirovi naziv se koristi
    izravno u IMAP SELECT (već je u ispravnom kodiranju)."""
    status, data = imap.list()
    if status != "OK":
        return []

    found = []
    for line in data:
        text = line.decode("utf-8", errors="replace")
        m = FOLDER_LIST_RE.match(text)
        if not m:
            continue
        raw_name = m.group("name")
        decoded_name = imap_utf7_decode(raw_name)

        for root in roots:
            if decoded_name == root or decoded_name.startswith(root + "/"):
                location = decoded_name.split("/")[0]
                found.append((raw_name, location))
                break
    return found


def find_course_code(identifier_line: str, course_codes: list) -> str:
    for code in sorted(course_codes, key=len, reverse=True):
        if code.lower() in identifier_line.lower():
            return code
    return ""


def build_search_query(config: dict) -> str:
    """IMAP SEARCH kriterij - mailovi od sender_filter adrese, opcionalno
    samo od since_date nadalje (npr. da se ignoriraju stare, prošlogodišnje
    prijave)."""
    criteria = f'FROM "{config["sender_filter"]}"'
    since_date = config.get("since_date")
    if since_date:
        imap_date = datetime.strptime(since_date, "%Y-%m-%d").strftime("%d-%b-%Y")
        criteria += f' SINCE "{imap_date}"'
    return f"({criteria})"


# --- Automatski odgovor (SMTP) ----------------------------------------------


def send_confirmation_email(config: dict, data: dict, course_code: str,
                             location: str, dates: str) -> bool:
    """Pošalje novu potvrdnu poruku prijavljenom polazniku (koristi istu
    app-lozinku kao IMAP, preko SMTP-a). Ovo je nova poruka njemu/njoj, a
    ne odgovor unutar niti izvorne obavijesti (ta obavijest ide samo tebi,
    prijavljeni nije primatelj tog maila pa nema na što nastaviti thread)."""
    to_email = (data.get("Email Address") or "").strip()
    if not to_email:
        return False

    template_vars = {
        "first_name": data.get("First name", ""),
        "last_name": data.get("Last Name", ""),
        "course_code": course_code,
        "location": location,
        "dates": dates,
        "instructor_name": config["instructor_name"],
    }
    subject = config["reply_subject"].format(**template_vars)
    body = config["reply_body"].format(**template_vars)

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = config["zoho_email"]
    msg["To"] = to_email
    msg.set_content(body)

    with smtplib.SMTP_SSL(config["smtp_host"], config.get("smtp_port", 465)) as smtp:
        smtp.login(config["zoho_email"], config["zoho_app_password"])
        smtp.send_message(msg)
    return True


# --- Građenje / popunjavanje Excel predloška --------------------------------


TEMPLATE_PATH = Path(__file__).with_name("template_admin_sheet.xlsx")


def build_course_workbook(course_code: str, location: str, dates: str,
                           instructor: str, currency: str = "eur") -> Workbook:
    """Učita pravi Emmett predložak (template_admin_sheet.xlsx) - ista
    vizualna forma, fontovi, boje, obrubi, visine redova kao original -
    i samo popuni metapodatke. Tablica polaznika i formule su već u
    predlošku, netaknuti."""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["podaci"]

    ws["C4"] = course_code
    ws["C5"] = location
    ws["C6"] = dates
    ws["M4"] = instructor
    ws["M6"] = currency

    return wb


def write_totals_formulas(ws, totals_row: int) -> None:
    last_row = totals_row - 1
    ws.cell(row=totals_row, column=14,
            value=f"=SUM(L{FIRST_PARTICIPANT_ROW}:L{last_row})").number_format = "0.00"

    ws.cell(row=totals_row + 1, column=10, value="Total Income")
    c = ws.cell(row=totals_row + 1, column=12,
                value=f"=SUM(L{FIRST_PARTICIPANT_ROW}:L{totals_row})")
    c.number_format = '#,##0.00\\ "€"'

    ws.cell(row=totals_row + 2, column=10, value="VAT")
    ws.cell(row=totals_row + 2, column=11, value=0)
    c = ws.cell(row=totals_row + 2, column=12, value=f"=L{totals_row + 1}/119*K{totals_row + 2}")
    c.number_format = '#,##0.00\\ "€"'

    ws.cell(row=totals_row + 3, column=10, value="Income minus VAT x%")
    c = ws.cell(row=totals_row + 3, column=12, value=f"=L{totals_row + 1}-L{totals_row + 2}")
    c.number_format = '#,##0.00\\ "€"'

    ws.cell(row=totals_row + 4, column=10, value="Commission Emmett (15 %)")
    c = ws.cell(row=totals_row + 4, column=12, value=f"=L{totals_row + 3}*0.15")
    c.number_format = '#,##0.00\\ "€"'

    ws.cell(row=totals_row + 5, column=10, value="Commission Bord (5%)")
    c = ws.cell(row=totals_row + 5, column=12, value=f"=L{totals_row + 3}*0.05")
    c.number_format = '#,##0.00\\ "€"'

    ws.cell(row=totals_row + 6, column=10, value="Commission Rep. (5%)")
    c = ws.cell(row=totals_row + 6, column=12, value=f"=L{totals_row + 3}*0.05")
    c.number_format = '#,##0.00\\ "€"'

    ws.cell(row=totals_row + 7, column=10, value="Income ")
    c = ws.cell(row=totals_row + 7, column=12,
                value=f"=L{totals_row + 3}-L{totals_row + 4}-L{totals_row + 5}-L{totals_row + 6}")
    c.number_format = '#,##0.00\\ "€"'


def find_totals_row(ws) -> int:
    for row in range(FIRST_PARTICIPANT_ROW, ws.max_row + 2):
        if ws.cell(row=row, column=10).value == "Total Income":
            return row - 1
    return TOTALS_ROW


def find_duplicate_row(ws, email: str, totals_row: int) -> int:
    """Vrati broj retka postojećeg polaznika s istom email adresom, ili
    None ako nema. Koristi se za detekciju duplih mailova (Emmettov
    sustav za prijave zna povremeno poslati istu prijavu dvaput)."""
    email_norm = (email or "").strip().lower()
    if not email_norm:
        return None
    for row in range(FIRST_PARTICIPANT_ROW, totals_row):
        existing = ws.cell(row=row, column=8).value  # Email Address
        if existing and str(existing).strip().lower() == email_norm:
            return row
    return None


def next_participant_row(ws, totals_row: int) -> int:
    for row in range(FIRST_PARTICIPANT_ROW, totals_row):
        if ws.cell(row=row, column=2).value in (None, ""):
            return row
    return None  # nema slobodnog reda, treba proširiti tablicu


def expand_participant_table(ws, totals_row: int) -> int:
    """Umetne jedan dodatni red prije retka s ukupnim iznosima i vrati
    indeks novog (slobodnog) reda za polaznika."""
    vat_rate = ws.cell(row=totals_row + 2, column=11).value  # sačuvaj ručno uneseni PDV %

    ws.insert_rows(totals_row)
    new_row = totals_row
    last_number = ws.cell(row=new_row - 1, column=1).value or (new_row - FIRST_PARTICIPANT_ROW)
    ws.cell(row=new_row, column=1, value=last_number + 1)
    ws.cell(row=new_row, column=9).number_format = PHONE_FORMAT
    ws.cell(row=new_row, column=12).number_format = '#,##0.00\\ "€"'

    new_totals_row = find_totals_row(ws)
    write_totals_formulas(ws, new_totals_row)
    if vat_rate is not None:
        ws.cell(row=new_totals_row + 2, column=11, value=vat_rate)
    return new_row


def append_participant(ws, data: dict, totals_row: int) -> int:
    if find_duplicate_row(ws, data.get("Email Address"), totals_row) is not None:
        return None  # duplikat - već postoji polaznik s tom email adresom

    row = next_participant_row(ws, totals_row)
    if row is None:
        row = expand_participant_table(ws, totals_row)

    ws.cell(row=row, column=2, value=data["First name"])
    ws.cell(row=row, column=3, value=data["Last Name"])
    ws.cell(row=row, column=4, value=data["Street "])
    ws.cell(row=row, column=5, value=data["City/ Town "])
    postal = data["Post/Zip Code "]
    ws.cell(row=row, column=7, value=int(postal) if postal.isdigit() else postal)
    ws.cell(row=row, column=8, value=data["Email Address"])
    ws.cell(row=row, column=9, value=data["Mobile Ph"])
    ws.cell(row=row, column=9).number_format = PHONE_FORMAT
    ws.cell(row=row, column=10, value="Croatia")
    ws.cell(row=row, column=11, value="N")
    return row


def get_or_create_workbook(path: Path, course_code: str, location: str,
                            dates: str, instructor: str):
    if path.exists():
        wb = load_workbook(path)
        ws = wb["podaci"]
        if dates and ws["C6"].value != dates:
            ws["C6"] = dates
        return wb, ws
    wb = build_course_workbook(course_code, location, dates, instructor)
    return wb, wb["podaci"]


def process_folder(imap, raw_folder: str, folder_location: str, config: dict,
                    processed: set, workbooks: dict) -> int:
    """Obradi jedan IMAP folder. folder_location je 'Split'/'Zagreb' (grad
    poznat iz naziva foldera) ili None (za INBOX - grad se mora pročitati
    iz teksta maila). Vraća broj dodanih prijava iz ovog foldera."""
    status, _ = imap.select(f'"{raw_folder}"')
    if status != "OK":
        print(f"[!] Ne mogu otvoriti folder {raw_folder!r}, preskačem.")
        return 0

    status, uid_data = imap.uid("search", None, build_search_query(config))
    if status != "OK" or not uid_data or uid_data[0] is None:
        return 0

    uids = uid_data[0].split()
    state_key_prefix = f"{raw_folder}::"
    new_uids = [
        uid.decode() for uid in uids
        if f"{state_key_prefix}{uid.decode()}" not in processed
    ]

    if not new_uids:
        return 0

    added = 0
    for uid in new_uids:
        state_key = f"{state_key_prefix}{uid}"

        status, msg_data = imap.uid("fetch", uid, "(RFC822)")
        if status != "OK" or not msg_data or msg_data[0] is None:
            continue

        raw = msg_data[0][1]
        msg = email.message_from_bytes(raw, policy=policy.default)
        text = get_plain_text(msg)
        parsed = parse_application(text)

        if folder_location:
            course_code = find_course_code(parsed["identifier_line"], config["course_codes"])
            location = folder_location if course_code else None
        else:
            course_code, location = match_course_and_location(
                parsed["identifier_line"], config["course_codes"],
            )

        if course_code and location:
            dates = extract_dates(parsed["identifier_line"], course_code)
            trailing_name = extract_trailing_name(parsed["identifier_line"], course_code)
            instructor = trailing_name or config["instructor_name"]

            key = (course_code, location)
            if key not in workbooks:
                filename = sanitize_filename(f"{course_code} {location}") + ".xlsx"
                path = Path(config["output_dir"]) / filename
                wb, ws = get_or_create_workbook(
                    path, course_code, location, dates, instructor
                )
                totals_row = find_totals_row(ws)
                workbooks[key] = [path, wb, ws, totals_row]
            else:
                path, wb, ws, totals_row = workbooks[key]
                if dates and ws["C6"].value != dates:
                    ws["C6"] = dates

            path, wb, ws, totals_row = workbooks[key]
            row = append_participant(ws, parsed, totals_row)

            if row is None:
                print(f"Preskočeno [{raw_folder}] - duplikat (email već postoji u "
                      f"{course_code} / {location}): {parsed['Email Address']}")
            else:
                if row >= totals_row:
                    totals_row = find_totals_row(ws)
                    workbooks[key][3] = totals_row
                added += 1
                print(f"Dodano [{raw_folder}] ({course_code} / {location}): "
                      f"{parsed['First name']} {parsed['Last Name']}")

                if config.get("send_replies"):
                    try:
                        sent = send_confirmation_email(
                            config, parsed, course_code, location,
                            ws["C6"].value or dates,
                        )
                        if sent:
                            print(f"  Poslana potvrda na {parsed['Email Address']}")
                        else:
                            print("  Potvrda nije poslana (nema email adrese)")
                    except Exception as e:
                        print(f"  Greška pri slanju potvrde: {e}")
        else:
            print(f"Preskočeno [{raw_folder}]: {parsed['identifier_line'][:80]}")

        processed.add(state_key)

    return added


def main():
    config = load_config()
    state_path = Path(config.get("state_path", "processed_uids.json"))
    output_dir = Path(config["output_dir"])
    output_dir.mkdir(parents=True, exist_ok=True)

    processed = load_state(state_path)

    imap = imaplib.IMAP4_SSL(config["imap_host"])
    imap.login(config["zoho_email"], config["zoho_app_password"])

    folder_roots = config.get("folder_roots", ["Split", "Zagreb"])
    target_folders = [("INBOX", None)] + discover_target_folders(imap, folder_roots)

    workbooks = {}  # (course_code, location) -> [path, wb, ws, totals_row]
    added = 0

    for raw_folder, folder_location in target_folders:
        added += process_folder(imap, raw_folder, folder_location, config, processed, workbooks)

    for path, wb, ws, _ in workbooks.values():
        wb.save(path)

    save_state(state_path, processed)
    imap.logout()

    if added == 0 and not workbooks:
        print("Nema novih mailova.")
    print(f"\nGotovo. Dodano {added} novih prijava u {len(workbooks)} datoteka "
          f"(pretraženo {len(target_folders)} foldera).")


if __name__ == "__main__":
    main()
